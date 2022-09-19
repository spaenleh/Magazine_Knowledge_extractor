import os
import re
from collections import Counter

import pandas as pd
import locale
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


import utils
import fitz

locale.setlocale(locale.LC_TIME, 'fr_FR')


def filter_out(entry, mag_name):
    if mag_name in utils.rules:
        mag_rules = utils.rules[mag_name]
    else:
        mag_rules = utils.rules[utils.DEFAULT]
    clean = []
    for rule in mag_rules:
        matched = re.search(rule, entry)
        clean.append(matched is None)
    return all(clean)


def get_info(file, mag_name):
    if mag_name in utils.titles:
        regex = utils.titles[mag_name]
    else:
        return {}
    match = re.match(regex, file)
    infos = {
        utils.MAG: match.group(utils.MAG),
        utils.DATE: match.group(utils.DATE).replace('-', ' '),
        utils.ISSUE_NUMBER: int(match.group(utils.ISSUE_NUMBER)),
        utils.TITLE: match.group(utils.TITLE).replace('-', ' '),
    }
    return infos


def should_keep(level, mag_name):
    if mag_name in utils.levels:
        keep_levels = utils.levels[mag_name]
    else:
        return False
    if level in keep_levels:
        return True
    else:
        return False


def export_to_excel(df, file_name):
    wb = Workbook()

    df_grouped = df.groupby(utils.MAG)
    for mag_name, small_df in df_grouped:
        ws = wb.create_sheet(mag_name)
        small_df.sort_values(utils.ISSUE_NUMBER, inplace=True)
        rows = dataframe_to_rows(small_df)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row[1:], 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file_name)
    ws = wb.active
    ws.title = "Sommaire"
    ws['A1'].value = 'Dernière mise à jour'
    ws['B1'].value = datetime.now().strftime("%d %B %Y")

    wb.save(file_name)


def main():
    records = []
    for root, subdirectories, files in os.walk('documents'):
        # print(root, subdirectories, files)
        mag_name = os.path.basename(root)
        for file in files:
            if not file.endswith('.pdf'):
                continue
            filename = os.path.join(root, file)
            # print(filename)
            infos = get_info(file, mag_name)
            # print(infos)
            # print(f"Processing: {filename}")
            doc = fitz.Document(filename)
            toc = doc.get_toc()
            # remove tags that are inappropriate
            toc = [entry for entry in toc if filter_out(entry[1], mag_name)]
            for item in toc:
                level, title, page_number = tuple(item)
                # in certain mags, only certain levels are of interest
                if not should_keep(level, mag_name):
                    continue
                title = title.replace('\xa0', ' ').replace('\n', '')
                # print(f"{title} (p{page_number})")
                record = infos.copy()
                record[utils.ARTICLE] = title
                record[utils.PAGE] = page_number
                records.append(record)

    df = pd.DataFrame.from_records(records)
    article_counter = sorted([(a, v) for a, v in Counter(df[utils.ARTICLE]).items()], key=lambda x: x[1], reverse=True)
    for item in article_counter[:10]:
        print(item)
    df = df[[utils.ARTICLE, utils.PAGE, utils.ISSUE_NUMBER, utils.DATE, utils.TITLE, utils.MAG]]
    df.to_csv('Index_articles.csv')
    export_to_excel(df, 'Index_articles.xlsx')


if __name__ == '__main__':
    main()
